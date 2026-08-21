"""Extract curated English place content from the supplied Champagne Atlas workbook."""

from __future__ import annotations

import hashlib
import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook


GRAPE_PATTERN = re.compile(r"^(.*?):\s*([0-9.]+)\s*ha;\s*([0-9.]+)%$")


def cell_text(value: object) -> str:
    return "" if value is None else str(value).strip()


def parse_grapes(value: object, place_id: str) -> list[dict[str, object]]:
    if cell_text(value).upper() in {"", "N/A", "NA", "NOT AVAILABLE"}:
        return []
    result = []
    for part in cell_text(value).split("|"):
        part = part.strip()
        if not part:
            continue
        match = GRAPE_PATTERN.match(part)
        if not match:
            raise ValueError(f"Invalid grape variety value for {place_id}: {part}")
        result.append(
            {
                "name": match.group(1).strip(),
                "hectares": float(match.group(2)),
                "percentage": float(match.group(3)),
            }
        )
    return result


def main() -> None:
    if len(sys.argv) != 3:
        raise SystemExit("Usage: extract-place-english-workbook.py INPUT.xlsx OUTPUT.json")
    source = Path(sys.argv[1]).resolve()
    output = Path(sys.argv[2]).resolve()
    workbook = load_workbook(source, data_only=True, read_only=True)

    place_sheet = workbook["place_records_202608172041"]
    rows = list(place_sheet.iter_rows(values_only=True))
    headers = [cell_text(value) for value in rows[0]]
    expected = [
        "Places",
        "Population",
        "Vineyard area",
        "Main grape",
        "Grand/Premier Cru",
        "Soil",
        "Wine character",
        "Grape varieties",
    ]
    if headers != expected:
        raise ValueError(f"Unexpected place headers: {headers}")

    sources_sheet = workbook["Sources"]
    source_rows = list(sources_sheet.iter_rows(values_only=True))
    source_headers = [cell_text(value) for value in source_rows[0]]
    if source_headers != ["Place", "Vineyard source", "Population source", "Cru source", "Note"]:
        raise ValueError(f"Unexpected source headers: {source_headers}")
    sources = {
        cell_text(row[0]): {
            "vineyardUrl": cell_text(row[1]),
            "populationUrl": cell_text(row[2]),
            "cruUrl": cell_text(row[3]),
            "note": cell_text(row[4]),
        }
        for row in source_rows[1:]
        if cell_text(row[0])
    }

    items = []
    for row in rows[1:]:
        place_id = cell_text(row[0])
        if not place_id:
            continue
        if place_id not in sources:
            raise ValueError(f"Missing source row for {place_id}")
        items.append(
            {
                "id": place_id,
                "soil": cell_text(row[5]),
                "wineCharacter": cell_text(row[6]),
                "grapeVarieties": parse_grapes(row[7], place_id),
                "sources": sources[place_id],
            }
        )

    ids = [item["id"] for item in items]
    if len(items) != 83 or len(set(ids)) != 83:
        raise ValueError(f"Expected 83 unique places, got {len(items)} rows and {len(set(ids))} IDs")

    payload = {
        "version": 1,
        "sourceWorkbook": source.name,
        "sourceWorkbookSha256": hashlib.sha256(source.read_bytes()).hexdigest(),
        "importedAt": "2026-08-20",
        "items": items,
    }
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"PLACE_EN_EXTRACT total={len(items)} output={output}")


if __name__ == "__main__":
    main()
